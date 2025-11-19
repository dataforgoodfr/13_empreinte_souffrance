import argparse
import json
import os
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# =============================================================================
# CONFIGURATION
# =============================================================================


class Config:
    # data folder name
    DATA_PATH = Path("data")

    # csv where processed product data is stored, several options
    INPUT_CSV_FILE = DATA_PATH / "processed_products.csv"
    INPUT_CSV_FILE_FR = DATA_PATH / "processed_products_fr.csv"
    INPUT_CSV_FILE_POTENTIAL = DATA_PATH / "processed_potential_eggs.csv"

    # excel output is formatted products_{dataset}_{subset}.xlsx
    OUTPUT_FILE_PREFIX = "products"

    DATASET_MAP = {
        "france": INPUT_CSV_FILE_FR,
        "potential": INPUT_CSV_FILE_POTENTIAL,
        "world": INPUT_CSV_FILE,
    }

    DEFAULT_SHEET_NAME = "Feuille1"
    EXCEL_EXTENSION = ".xlsx"

    # Images configuration
    IMAGE_COUNT = 10
    ROW_HEIGHT = 300
    HEADER_FONT_SIZE = 8
    TEST_SAMPLE_SIZE = 30

    ROW_FREEZE = 1
    COL_FREEZE = 5

    # To generate product url
    PRODUCT_BASE_URL = "https://fr.openfoodfacts.org/produit/"

    # to generate image urls
    IMAGE_BASE_URL = "https://images.openfoodfacts.net/images/products/"

    # Base columns now always include breeding + quantity fields
    BASE_COLUMNS = [
        "Code et URL produit",
        "Elevage",
        "Quantité",
        "Calibre",
        "Catégories",
        "Autres",
    ]

    ANALYSIS_COLUMNS = [
        "OCR",
        "breeding_type_related",
        "weight_related",
        "predictions_categories",
    ]

    # How column names are displayed
    COLUMN_RENAMES: Dict[str, str] = {
        "Elevage": "Elevage (3, 2, 1, 0)",
        "Quantité": "Quantité (nb d'œufs ou poids)",
        "Calibre": "Calibre (S, M, L, XL)",
        "predictions_categories": "predictions categories",
        "Image front": "Image principale",
        "Image ingredients": "Image ingrédients",
        "Autres": "Nom, labels, ingrédients",
    }

    COLUMN_WIDTHS = {
        "Code et URL produit": 14,
        "Autres": 22,
        "Catégories": 22,
        "Quantité": 15,
        "Elevage": 15,
        "Calibre": 17,
        "OCR": 50,
        "breeding_type_related": 15,
        "weight_related": 15,
        "predictions_categories": 15,
    }

    # add fixed width for all image columns
    @classmethod
    def column_widths(cls):
        return {
            **{cls.COLUMN_RENAMES.get(k, k): v for k, v in cls.COLUMN_WIDTHS.items()},
            **{f"Image {i}": 50 for i in range(1, cls.IMAGE_COUNT + 1)},
        }


# =============================================================================
# UTILITIES
# =============================================================================


def safe_json_load(raw: Any) -> Any:
    """Load json formatted fields from dataframe for multilingual purpose"""
    if isinstance(raw, str):
        try:
            cleaned = raw.replace("'", '"').replace("None", "null")
            return json.loads(cleaned)
        except json.JSONDecodeError:
            return raw
    return raw


def load_product_data(file_path: str) -> pd.DataFrame:
    """Load csv file from path and return dataframe"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Data file not found: {file_path}")
    df = pd.read_csv(file_path)
    if df.empty:
        raise pd.errors.EmptyDataError(f"File is empty: {file_path}")
    print(f"Loaded {len(df)} products from {file_path}")
    return df


def get_value(df: pd.DataFrame, code: str, column: str) -> Optional[str]:
    """Get a column value for a given dataframe, barcode and column name"""
    mask = df["code"] == code
    if not mask.any():
        return None
    try:
        value = df.loc[mask, column].iloc[0]
    except (KeyError, IndexError):
        return None
    return str(value) if pd.notna(value) else None


def fetch_product(df: pd.DataFrame, code: str) -> Dict:
    row = df[df["code"] == code]
    return row.iloc[0].replace({np.nan: None}).to_dict() if not row.empty else {}


# =============================================================================
# MISSING-DATA RESTRICTION
# =============================================================================


def get_product_codes_missing_data(df: pd.DataFrame) -> List[str]:
    """Return product codes with missing breeding type or quantity data."""
    mask_breeding_missing = df["breeding"].isin(["not managed", "no breeding"]) | df["breeding"].isna()
    mask_quantity_missing = (df["egg_count"] < 0) | (df["egg_count"].isna()) | (df["egg_count"] == 0)
    return df.loc[mask_breeding_missing | mask_quantity_missing, "code"].tolist()


# =============================================================================
# DATA PARSING
# =============================================================================


def parse_multilingual(field: Any) -> str:
    """
    Parse multilingual fields from OpenFoodFacts format to a single string and gets all languages
    Excludes main to avoid redundancy.
    """
    data = safe_json_load(field)
    if isinstance(data, list):
        return "\n".join(d.get("text", "") for d in data if isinstance(d, dict) and d.get("lang") != "main")
    return str(data) if data else ""


def build_image_urls(code: str, images_raw: Any) -> List[str]:
    """Build image_urls from product code and follow OpenFoodFacts path scheme"""
    images = safe_json_load(images_raw) or []
    code_str = str(code).zfill(13)
    path = f"{code_str[:3]}/{code_str[3:6]}/{code_str[6:9]}/{code_str[9:]}"
    keys = sorted(
        [img["key"] for img in images if isinstance(img, dict) and str(img.get("key")).isdigit()],
        key=int,
    )
    return [f"{Config.IMAGE_BASE_URL}{path}/{k}.jpg" for k in keys][: Config.IMAGE_COUNT]


def get_analysis_fields(df: pd.DataFrame, code: str) -> Dict[str, str]:
    """Get analysis-related fields from dataframe for a given product code"""
    proba = [v for v in (get_value(df, code, col) for col in ["proba_1", "proba_2", "proba_3"]) if v is not None]
    return {
        "OCR": get_value(df, code, "ocr_text") or "",
        "breeding_type_related": get_value(df, code, "breeding_type_related") or "",
        "weight_related": get_value(df, code, "weight_related") or "",
        "predictions_categories": "\n".join(proba),
    }


# =============================================================================
# ROW BUILDER
# =============================================================================


def build_product_row(product: Dict, code: str, df: pd.DataFrame) -> Dict[str, Any]:
    """Define the content of each product row by parsing product data and associated dataframe"""

    name = parse_multilingual(product.get("product_name", []))
    generic = parse_multilingual(product.get("generic_name", ""))

    ingredients = product.get("ingredients_tags", [])
    labels = product.get("labels_tags", [])
    categories_raw = safe_json_load(product.get("categories_tags", ""))
    categories = [c for c in (categories_raw or []) if c]
    breeding = get_value(df, code, "breeding") or ""
    quantity = get_value(df, code, "egg_count") or ""
    caliber = get_value(df, code, "caliber") or ""

    row = {
        "Code et URL produit": code,
        "Autres": f"{name}\n{generic}\n{ingredients}\n{labels}",
        "Catégories": "\n".join(categories),
        "Elevage": breeding,
        "Quantité": quantity,
        "Calibre": caliber,
    }

    images = build_image_urls(code, product.get("images", []))
    for idx in range(Config.IMAGE_COUNT):
        row[f"Image {idx + 1}"] = f'=IMAGE("{images[idx]}")' if idx < len(images) else ""

    row.update(get_analysis_fields(df, code))
    return row


def column_order() -> List[str]:
    """Define column order from configuration lists"""
    cols = Config.BASE_COLUMNS
    cols += Config.ANALYSIS_COLUMNS
    cols += [f"Image {i + 1}" for i in range(Config.IMAGE_COUNT)]
    return cols


# =============================================================================
# EXCEL OUTPUT
# =============================================================================


def apply_hyperlinks(ws, columns, product_codes):
    """
    Apply hyperlinks to the "Code et URL produit" column for each product code.
    """
    if "Code et URL produit" not in columns:
        return

    code_col = columns.index("Code et URL produit") + 1
    for r_idx, code in enumerate(product_codes, start=2):
        cell = ws.cell(row=r_idx, column=code_col)
        cell.hyperlink = f"{Config.PRODUCT_BASE_URL}{code}"
        cell.style = "Hyperlink"


def apply_styles(ws, columns, product_codes):
    """
    Apply formatting and styles to the worksheet and optionally add hyperlinks.
    """
    border = Border(
        left=Side(style="dotted", color="E0E0E0"),
        right=Side(style="dotted", color="E0E0E0"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    apply_hyperlinks(ws, columns, product_codes)

    # Header
    for col_idx, name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = Config.column_widths().get(name, 15)
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(size=Config.HEADER_FONT_SIZE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Body
    for r_idx, code in enumerate(product_codes, start=2):
        ws.row_dimensions[r_idx].height = Config.ROW_HEIGHT
        for c_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
            if col_name in ["OCR", "Code et URL produit"]:
                cell.font = Font(size=8)

    ws.freeze_panes = ws.cell(row=Config.ROW_FREEZE + 1, column=Config.COL_FREEZE + 1)


def write_and_format_excel(df: pd.DataFrame, file_path: Path, product_codes: List[str]):
    """Remove existing file and creates a new one from dataframe"""
    wb = Workbook()
    ws = wb.active
    ws.title = Config.DEFAULT_SHEET_NAME

    # Write header
    for col_idx, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col)

    # Write rows
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    apply_styles(ws, df.columns.tolist(), product_codes)
    wb.save(file_path)


def generate_excel(
    df: pd.DataFrame,
    output: Path,
    product_codes: List[str] = [],
    show_progress: bool = True,
) -> Path:
    rows = []
    valid_codes = []
    """
    Generate an Excel file containing product data for the specified product codes
    from the DataFrame.
    """

    # If no product codes provided, use all codes from the DataFrame
    product_codes = df["code"].tolist() if product_codes == [] else product_codes

    # Iterate over product codes with progress bar to build product rows
    iterator = tqdm(product_codes, desc="Processing", unit="product") if show_progress else product_codes
    for code in iterator:
        product = fetch_product(df, code)
        if not product:
            continue
        row = build_product_row(product, code, df)
        rows.append(row)
        valid_codes.append(code)

    if not rows:
        raise ValueError("No valid products to process")

    df_out = pd.DataFrame(rows)

    # Keep only columns present in df_out according to predefined column_order()
    ordered = [c for c in column_order() if c in df_out.columns]
    df_out = df_out.reindex(columns=ordered + [c for c in df_out.columns if c not in ordered])
    df_out.rename(columns=Config.COLUMN_RENAMES, inplace=True)

    write_and_format_excel(df_out, output, valid_codes)

    return output


# =============================================================================
# HIGH-LEVEL COMMANDS
# =============================================================================


def create_test_excel(df: pd.DataFrame, output: Path) -> None:
    """Create an Excel file containing a random sample of products from the DataFrame."""
    sample_df = df.sample(n=min(len(df), Config.TEST_SAMPLE_SIZE))
    generate_excel(df=sample_df, output=output, show_progress=False)
    if os.name == "nt":
        os.startfile(output)


def create_all_products_excel(df: pd.DataFrame, output: Path) -> None:
    """Create an Excel file containing all products from the DataFrame."""
    generate_excel(df=df, output=output)


def create_missing_data_excel(df: pd.DataFrame, output: Path) -> None:
    """
    Create an Excel file containing products with missing breeding type
    or quantity data from the DataFrame.
    """
    missing_codes = get_product_codes_missing_data(df)
    generate_excel(product_codes=missing_codes, df=df, output=output)


# =============================================================================
# CLI
# =============================================================================


def parse_arguments() -> argparse.Namespace:
    """
    Parse command-line arguments for dataset and subset selection.
    Parsed arguments include:
        --dataset: choice of dataset among 'world', 'france', 'potential' (default: 'france')
        --subset: choice of subset among 'all', 'test', 'missing-data' (default: 'all')
    """
    parser = argparse.ArgumentParser(description="Export product data to Excel")

    parser.add_argument(
        "--dataset",
        choices=["world", "france", "potential"],
        default="france",
    )

    parser.add_argument(
        "--subset",
        choices=["all", "test", "missing-data"],
        default="all",
    )

    return parser.parse_args()


def main():
    """
    Main function to execute the script based on command-line arguments.
    loads the specified dataset, defines the output file name
    and generates an Excel file with the results for the specified subset.
    """
    args = parse_arguments()

    file_path = Config.DATASET_MAP[args.dataset]
    df = load_product_data(file_path)

    output = Config.DATA_PATH / f"{Config.OUTPUT_FILE_PREFIX}_{args.dataset}_{args.subset}.xlsx"

    if args.subset == "test":
        create_test_excel(df, output)
    elif args.subset == "missing-data":
        create_missing_data_excel(df, output)
    else:
        create_all_products_excel(df=df, output=output)


if __name__ == "__main__":
    main()
